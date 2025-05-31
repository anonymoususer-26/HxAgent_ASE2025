from core.short_term_memory import ShortTermMemory
from core.long_term_memory import LongTermMemory
from dataset.task_factory import Factory, MiniwobTaskFatory
from core.agent_access import AgentAccess
from core.action import InputAction, ClickAction, EnterAction, ChainAction
from core.utils import remove_whitespace, remove_special_characters, extract_steps, find_best_matching_action
from core.logger import ConsoleLogger
from core.simulator import Simulator
from core.storage import Storage
from config import GLOBAL_CONFIG
from core.state import State
import statistics
import traceback
import openpyxl
import time
import os
from enum import StrEnum
from core.agent import APIUsageManager
from pyvis.network import Network
import uuid
from selenium.webdriver.common.by import By
import pandas as pd
import json

simulator = Simulator()
has_set_up = False

class Session:
    def __init__(self, task_factory: Factory):
        self.store = Storage()
        self.task_factory = task_factory
        self.long_term_memory = LongTermMemory()
        if not os.path.exists('log/session'):
            os.makedirs('log/session')
        global has_set_up
        if not has_set_up:
            ConsoleLogger.instruction(f'Please set up the browser for the experiment, this includes logging in to necessary sites. Enter any key when you finish.')
            _ = input()
            has_set_up = True

    def train(self, total_iteration = 25, max_steps = 20):
        global simulator
        ConsoleLogger.info(f'Training started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'Failed reason', 'Token usage'])
        experience = ""
        score_history = []
        best_moving_average = 0
        should_terminate_early = False

        for iteration in range(total_iteration):
            APIUsageManager.reset()
            if (should_terminate_early):
                ConsoleLogger.info(f'Training session terminating...')
                if (best_moving_average == 0):
                    self.store.save_optimal_knowledge(f"storage/{self.task_factory.id}.txt", self.long_term_memory.stringify())
                break

            ConsoleLogger.info(f'Training iteration {iteration + 1} started')
            is_successful = None
            failed_reason = '-'
            start_time = time.time()
            simulator.restart(self.task_factory.url)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {task}')
            ConsoleLogger.instruction(f'Do u want to skip this task: ')
            is_skip = False if input() == "0" else True
            if (is_skip):
                continue
            short_term_memory = ShortTermMemory(self.task_factory.url, self.task_factory.title, task)
            time.sleep(4)
       
            try: 
                for step in range(max_steps):
                    actions, action_mp = AgentAccess.generate_actions(simulator.get_clickables(), short_term_memory.stringify(), experience)
                    state = State(simulator.crawl(), simulator.get_clickables(), actions, simulator.get_screenshot_url())
                    state.try_describe_state(short_term_memory.stringify())

                    is_done = step != 0 and (simulator.has_reach_end() or state.is_done)
                    if (is_done or state.early_stop):
                        is_successful = simulator.is_successful()
                        if is_successful == None:
                            ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                            is_successful = False if input() == "0" else True
                        if (not is_successful):
                            ConsoleLogger.instruction(f'Please specify why this task ({task}) failed:')
                            failed_reason = input()
                            failed_reason = ""
                        self.long_term_memory.update(short_term_memory, is_successful, failed_reason)
                        ConsoleLogger.info(f'Current best moving average is', best_moving_average)
                        ConsoleLogger.info(f'Current score history is', score_history)
                        ConsoleLogger.instruction(f'Should training session be terminated now? (meaning the result is good enough)')
                        should_terminate_early = False if input() == "0" else True
                        break
                    
                    print("====================================")
                    print(state.description)
                    print("====================================")
                    
                    chosen_action, chosen_action_reason = AgentAccess.choose_action(actions, state.description, short_term_memory.stringify(), experience)

                    if ('duplicate' in chosen_action.description["target object"]):
                        if (chosen_action.description["target object"]["duplicate"]):
                            ConsoleLogger.info('Duplicate action detected, using vision...')
                            chosen_action_cp = chosen_action.description.copy()
                            chosen_action_cp["target object"].pop("duplicate")
                            chosen_action_cp["target object"].pop("duplicate_index")
                            duplicate_key = remove_special_characters(str(chosen_action_cp["operation"]) + str(chosen_action_cp["target object"]))
                            
                            duplicate_actions_idx = action_mp[duplicate_key]
                            # Get the list of duplicate actions from the action list
                            duplicate_actions = [actions[i] for i in duplicate_actions_idx]
                            for idx, action in enumerate(duplicate_actions):
                                elements = simulator.find_elements(By.XPATH, action.xpath)
                                simulator.highlight_element(elements[0], idx + 1, duration=5)

                            with open(f'log/screenshot/{self.task_factory.id}_{iteration}_{step}.png', 'wb+') as f:
                                f.write(simulator.get_screenshot())

                            # Using vision to choose the best action
                            chosen_action, chosen_action_reason = AgentAccess.choose_action(duplicate_actions, "", short_term_memory.stringify(), experience, simulator.get_screenshot_url())


                    ConsoleLogger.info(f'Chosen action: {chosen_action.description}, reason: {chosen_action_reason}')
                    try:
                        chosen_action.perform(simulator)
                        time.sleep(3)
                        short_term_memory.append(chosen_action, state)
                    except Exception as e:
                        ConsoleLogger.error(f'Ignored error encountered when perform action: {e}')
                        ConsoleLogger.instruction(f'Should we add this action to short-term-memory? (0 = no, 1 = yes):')
                        is_implementation_fault = False if input() == "0" else True
                        if (is_implementation_fault):
                            short_term_memory.append(chosen_action, state)
                    if step == max_steps - 1:
                        is_successful = False
                        failed_reason = f"Step overflow"
                        ConsoleLogger.error(f'Step overflow when solving task')
                        # ConsoleLogger.instruction(f'Please specify why this task ({task}) failed:')
                        # failed_reason = input()
                        self.long_term_memory.update(short_term_memory, is_successful, failed_reason)
            except Exception as e:
                is_successful = False
                failed_reason = f"Unexpected error: {e}"
                ConsoleLogger.error(f'Error encountered when solving task: {e}')
                traceback.print_exc()
            end_time = time.time()
            ws.append([iteration, task, short_term_memory.stringify(), int(is_successful), end_time - start_time, failed_reason, str(APIUsageManager.get_usage())])
            wb.save(f'log/session/{self.task_factory.id}_training.xlsx')
            experience = self.long_term_memory.stringify()
            score_history.append(int(is_successful))
            if (iteration + 1 >= GLOBAL_CONFIG['session']['training']['window_size']):
                moving_average = sum(score_history[-GLOBAL_CONFIG['session']['training']['window_size']:])/GLOBAL_CONFIG['session']['training']['window_size']
                if moving_average >= best_moving_average:
                    best_moving_average = moving_average
                    self.store.save_optimal_knowledge(f"storage/{self.task_factory.id}.txt", self.long_term_memory.stringify())
            ConsoleLogger.info(f'Training iteration {iteration + 1} completed')
        ConsoleLogger.info(f'Training ended for url {self.task_factory.url}')

    def evaluate(self, total_iteration = 25, max_steps = 15):
        global simulator
        ConsoleLogger.info(f'Evaluation started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'Failed Reason', 'Token usage', 'Test Script'])
        experience = self.store.load_optimal_knowledge(f'storage_claude/{self.task_factory.id}.txt')
        experience = ""
        count = 0
        for iteration in range(total_iteration):
            APIUsageManager.reset()
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} started')
            is_successful = None
            failed_reason = '-'
            start_time = time.time()
            simulator.restart(self.task_factory.url)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {task}')
            ConsoleLogger.instruction(f'Do u want to skip this task: ')
            is_skip = False if input() == "0" else True
            if (is_skip):
                continue
            short_term_memory = ShortTermMemory(self.task_factory.url, self.task_factory.title, task)
            try: 
                for step in range(max_steps):                        
                    is_done = step != 0 and (simulator.has_reach_end() or AgentAccess.is_done(short_term_memory.stringify()))
                    if (is_done):
                        is_successful = simulator.is_successful()
                        if is_successful == None:
                            ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                            is_successful = False if input() == "0" else True
                        if (not is_successful):
                            ConsoleLogger.instruction(f'Please specify why this task ({task}) failed:')
                            failed_reason = input()
                        break
                    actions, action_mp = AgentAccess.generate_actions(simulator.get_clickables(), short_term_memory.stringify(), experience)
                    state = State(simulator.crawl(), simulator.get_clickables(), actions, simulator.get_screenshot_url())
                    state.try_describe_state(short_term_memory.stringify())
                    print("====================================")
                    print(state.description)
                    print("====================================")
                    chosen_action, chosen_action_reason = AgentAccess.choose_action(actions, state.description, short_term_memory.stringify(), experience)
                    ConsoleLogger.info(f'Chosen action: {chosen_action.description}, reason: {chosen_action_reason}')
                    if ('duplicate' in chosen_action.description["target object"]):
                        if (chosen_action.description["target object"]["duplicate"]):
                            ConsoleLogger.info('Duplicate action detected, using vision...')
                            chosen_action_cp = chosen_action.description.copy()
                            chosen_action_cp["target object"].pop("duplicate")
                            chosen_action_cp["target object"].pop("duplicate_index")
                            duplicate_key = remove_special_characters(str(chosen_action_cp["operation"]) + str(chosen_action_cp["target object"]))
                            
                            duplicate_actions_idx = action_mp[duplicate_key]
                            # Get the list of duplicate actions from the action list
                            duplicate_actions = [actions[i] for i in duplicate_actions_idx]
                            for idx, action in enumerate(duplicate_actions):
                                elements = simulator.find_elements(By.XPATH, action.xpath)
                                simulator.highlight_element(elements[0], idx + 1, duration=5)

                            with open(f'log/screenshot/{self.task_factory.id}_{iteration}_{step}.png', 'wb+') as f:
                                f.write(simulator.get_screenshot())

                            # Using vision to choose the best action
                            chosen_action, chosen_action_reason = AgentAccess.choose_action(duplicate_actions, "", short_term_memory.stringify(), experience, simulator.get_screenshot_url())

                    try:
                        chosen_action.perform(simulator)
                        time.sleep(3)
                        short_term_memory.append(chosen_action, state)
                        print("====================================")
                        print(short_term_memory.stringify())
                        print("====================================")
                    except Exception as e:
                        ConsoleLogger.error(f'Ignored error encountered when perform action: {e}')
                        ConsoleLogger.instruction(f'Should we add this action to short-term-memory? (0 = no, 1 = yes):')
                        is_implementation_fault = False if input() == "0" else True
                        if (is_implementation_fault):
                            short_term_memory.append(chosen_action, state)
                    if step == max_steps - 1:
                        is_successful = False
                        failed_reason = f"Step overflow"
            except Exception as e:
                is_successful = False
                failed_reason = f"Unexpected error: {e}"
                ConsoleLogger.error(f'Error encountered when solving task: {e}')
                traceback.print_exc()
            end_time = time.time()
            ws.append([iteration, task, short_term_memory.stringify(), int(is_successful), end_time - start_time, failed_reason, str(APIUsageManager.get_usage()), short_term_memory.to_script(simulator)])
            wb.save(f'log/session_scripts/{self.task_factory.id}_evaluation.xlsx')
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} completed')
        ConsoleLogger.info(f'Evaluation ended for url {self.task_factory.url}')

    def playback(self):
        global simulator
        ConsoleLogger.info(f'Evaluation started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'Failed Reason', 'Token usage', 'Test Script'])
        directory = "log/session_miniwob"
        data = pd.read_excel(os.path.join(directory, f'{self.task_factory.id}_evaluation.xlsx'))
        
        for idx, row in data.iterrows():
            ConsoleLogger.info(f'Evaluation iteration {idx + 1} started')
            is_successful = None
            failed_reason = '-'
            start_time = time.time()
            simulator.restart(self.task_factory.url, row['Iteration'])
            time.sleep(3)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {row['Task']}')
            ConsoleLogger.instruction(f'Do u want to skip this task: ')
            is_skip = False if input() == "0" else True
            if (is_skip):
                continue
            short_term_memory = ShortTermMemory(self.task_factory.url, self.task_factory.title, row['Task'])

            path_text = row['Chosen Path']
            sequence_actions = extract_steps(path_text)
            try:
                for i, step in enumerate(sequence_actions, 1):
                    actions, action_mp = AgentAccess.generate_actions(simulator.get_clickables(), short_term_memory.stringify())
                    for action in actions:
                        action.description['target object'].pop('duplicate', None)
                        action.description['target object'].pop('duplicate_index', None)
                        print(action.description)
                    state = State(simulator.crawl(), simulator.get_clickables(), actions, simulator.get_screenshot_url())
                    matched_action = find_best_matching_action(step, actions)
                    if matched_action:
                        try:
                            print(f"\nStep {i}:")
                            print(f"Original step: {json.dumps(step, indent=2)}")
                            print(f"Matched action: {json.dumps(matched_action.description, indent=2)}")
                            if matched_action.description['operation'] == 'input':
                                matched_action.description['content'] = step['content']
                                matched_action = InputAction(matched_action.xpath, step['content'], matched_action.description)
                                if step["operation"] == "input + enter":
                                    enter_action = EnterAction(matched_action.xpath)
                                    matched_action = ChainAction([matched_action, enter_action])
                            
                            matched_action.perform(simulator)
                            if (i >= 13):
                                time.sleep(i)
                            else:
                                time.sleep(2)
                            short_term_memory.append(matched_action, state)
                            print("====================================")
                            print(short_term_memory.stringify())
                            print("====================================")
                        except Exception as e:
                            ConsoleLogger.error(f'Ignored error encountered when perform action: {e}')
                            ConsoleLogger.instruction(f'Should we add this action to short-term-memory? (0 = no, 1 = yes):')
                            is_implementation_fault = False if input() == "0" else True
                            if (is_implementation_fault):
                                short_term_memory.append(matched_action, state)
                    else:
                        print(f"\nStep {i}: No matching action found")

            except Exception as e:
                is_successful = False
                failed_reason = f"Unexpected error: {e}"
                ConsoleLogger.error(f'Error encountered when solving task: {e}')
                traceback.print_exc()
            end_time = time.time()
            ws.append([idx, row['Task'], short_term_memory.stringify(), row['Status'], row['Elapsed Time'], failed_reason, "", short_term_memory.to_script(simulator)])
            wb.save(f'log/session_scripts/{self.task_factory.id}_evaluation.xlsx')
            ConsoleLogger.info(f'Evaluation iteration {idx + 1} completed')
        ConsoleLogger.info(f'Evaluation ended for url {self.task_factory.url}')

class NoIterativePlanningSession:
    def __init__(self, task_factory: Factory):
        self.task_factory = task_factory
        if not os.path.exists('log/no_iterative_planning_session'):
            os.makedirs('log/no_iterative_planning_session')
        
        global has_set_up
        if not has_set_up:
            ConsoleLogger.instruction(f'Please set up the browser for the experiment, this includes logging in to necessary sites. Enter any key when you finish.')
            _ = input()
            has_set_up = True
    
    def train(self, total_iteration = 25, max_steps = 20):
        return

    def evaluate(self, total_iteration = 25, max_steps = 20):
        ConsoleLogger.info(f'Ablation study: no iterative, no short term, no experiment')

        global simulator
        ConsoleLogger.info(f'Evaluation started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'Failed Reason'])
        
        for iteration in range(total_iteration):
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} started')
            is_successful = None
            failed_reason = '-'
            start_time = time.time()
            simulator.restart(self.task_factory.url)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {task}')
            try:
                actions = AgentAccess.generate_actions(simulator.get_clickables(), "")
                chosen_path = AgentAccess.gen_plan(task, actions)
                ConsoleLogger.instruction(f'The agent generate the following plan: \n{chosen_path}\n')
                ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                is_successful = False if input() == "0" else True
                if (not is_successful):
                    ConsoleLogger.instruction(f'Please specify why this task ({task}) failed:')
                    failed_reason = input()
            except Exception as e:
                is_successful = False
                chosen_path = "-"
                failed_reason = f'Unexpected error: {e}'
                ConsoleLogger.error(f'Error encountered when solving task: {e}')
                traceback.print_exc()
            end_time = time.time()
            ws.append([iteration, task, chosen_path, int(is_successful), end_time - start_time, failed_reason])
            wb.save(f'log/no_iterative_planning_session/{self.task_factory.id}_evaluation.xlsx')
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} completed')
        ConsoleLogger.info(f'Evaluation ended for url {self.task_factory.url}')

class NoShortTermNoExpSession:
    def __init__(self, task_factory: Factory):
        self.task_factory = task_factory
        if not os.path.exists('log/no_short_term_no_exp_session'):
            os.makedirs('log/no_short_term_no_exp_session')
        global has_set_up
        if not has_set_up:
            ConsoleLogger.instruction(f'Please set up the browser for the experiment, this includes logging in to necessary sites. Enter any key when you finish.')
            _ = input()
            has_set_up = True
    
    def train(self, total_iteration = 25, max_steps = 20):
        return

    def evaluate(self, total_iteration = 25, max_steps = 20):
        ConsoleLogger.info(f'Ablation study: no short term, no experiment')

        global simulator
        ConsoleLogger.info(f'Evaluation started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'Failed Reason'])

        for iteration in range(total_iteration):
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} started')
            is_successful = None
            failed_reason = '-'
            start_time = time.time()
            simulator.restart(self.task_factory.url)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {task}')
            short_term_memory = ShortTermMemory(self.task_factory.url, self.task_factory.title, task)
            try: 
                for step in range(max_steps):
                    is_done = step != 0 and (simulator.has_reach_end() or AgentAccess.is_done(short_term_memory.stringify_short()))
                    if (is_done):
                        is_successful = simulator.is_successful()
                        if is_successful == None:
                            ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                            is_successful = False if input() == "0" else True
                        if (not is_successful):
                            ConsoleLogger.instruction(f'Please specify why this task ({task}) failed:')
                            failed_reason = input()
                        break
                    actions = AgentAccess.generate_actions(simulator.get_clickables(), "")
                    state = State(simulator.crawl(), simulator.get_clickables(), actions, simulator.get_screenshot_base64())
                    chosen_action, chosen_action_reason = AgentAccess.choose_action(actions, "", "")
                    ConsoleLogger.info(f'Chosen action: {chosen_action.description}, reason: {chosen_action_reason}')
                    try:
                        chosen_action.perform(simulator)
                        short_term_memory.append(chosen_action, state)
                    except Exception as e:
                        ConsoleLogger.error(f'Ignored error encountered when perform action: {e}')
                        ConsoleLogger.instruction(f'Should we add this action to short-term-memory? (0 = no, 1 = yes):')
                        is_implementation_fault = False if input() == "0" else True
                        if (is_implementation_fault):
                            short_term_memory.append(chosen_action, state)
                    if step == max_steps - 1:
                        is_successful = False
                        failed_reason = f"Step overflow"
            except Exception as e:
                is_successful = False
                failed_reason = f"Unexpected error: {e}"
                ConsoleLogger.error(f'Error encountered when solving task: {e}')
                traceback.print_exc()
            end_time = time.time()
            ws.append([iteration, task, short_term_memory.stringify_short(), int(is_successful), end_time - start_time, failed_reason])
            wb.save(f'log/no_short_term_no_exp_session/{self.task_factory.id}_evaluation.xlsx')
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} completed')
        ConsoleLogger.info(f'Evaluation ended for url {self.task_factory.url}') 

class NoExpSession:
    def __init__(self, task_factory: Factory):
        self.task_factory = task_factory
        if not os.path.exists('log/no_exp_session'):
            os.makedirs('log/no_exp_session')
        global has_set_up
        if not has_set_up:
            ConsoleLogger.instruction(f'Please set up the browser for the experiment, this includes logging in to necessary sites. Enter any key when you finish.')
            _ = input()
            has_set_up = True
    
    def train(self, total_iteration = 25, max_steps = 20):
        return

    def evaluate(self, total_iteration = 25, max_steps = 20):
        ConsoleLogger.info(f'Ablation study: no experience')

        global simulator
        ConsoleLogger.info(f'Evaluation started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'Failed Reason'])
        
        for iteration in range(total_iteration):
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} started')
            is_successful = None
            failed_reason = '-'
            start_time = time.time()
            simulator.restart(self.task_factory.url)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {task}')
            short_term_memory = ShortTermMemory(self.task_factory.url, self.task_factory.title, task)
            try: 
                for step in range(max_steps):
                    is_done = step != 0 and (simulator.has_reach_end() or AgentAccess.is_done(short_term_memory.stringify()))
                    if (is_done):
                        is_successful = simulator.is_successful()
                        if is_successful == None:
                            ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                            is_successful = False if input() == "0" else True
                        if (not is_successful):
                            ConsoleLogger.instruction(f'Please specify why this task ({task}) failed:')
                            failed_reason = input()
                        break
                    actions = AgentAccess.generate_actions(simulator.get_clickables(), short_term_memory.stringify())
                    state = State(simulator.crawl(), simulator.get_clickables(), actions, simulator.get_screenshot_base64())
                    state.try_describe_state(short_term_memory.stringify())
                    chosen_action, chosen_action_reason = AgentAccess.choose_action(actions, state.description, short_term_memory.stringify())
                    ConsoleLogger.info(f'Chosen action: {chosen_action.description}, reason: {chosen_action_reason}')
                    try:
                        chosen_action.perform(simulator)
                        short_term_memory.append(chosen_action, state)
                    except Exception as e:
                        ConsoleLogger.error(f'Ignored error encountered when perform action: {e}')
                        ConsoleLogger.instruction(f'Should we add this action to short-term-memory? (0 = no, 1 = yes):')
                        is_implementation_fault = False if input() == "0" else True
                        if (is_implementation_fault):
                            short_term_memory.append(chosen_action, state)
                    if step == max_steps - 1:
                        is_successful = False
                        failed_reason = f"Step overflow"
            except Exception as e:
                is_successful = False
                failed_reason = f"Unexpected error: {e}"
                ConsoleLogger.error(f'Error encountered when solving task: {e}')
            end_time = time.time()
            ws.append([iteration, task, short_term_memory.stringify(), int(is_successful), end_time - start_time, failed_reason])
            wb.save(f'log/no_exp_session/{self.task_factory.id}_evaluation.xlsx')
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} completed')
        ConsoleLogger.info(f'Evaluation ended for url {self.task_factory.url}')

class STATUS(StrEnum):
    FAILED = "However, your actions did not complete the goal. Now, you need to identify the earliest critical step where you made a mistake, and suggest a correction."
    CYCLE = "However, your actions led you to a loop that did not progress the task. Now, you need to identify the earliest critical step where you made a mistake, and suggest a correction."
    NO_CHANGE = "However, your last action did not cause anything to change on the last screen. You probably used the wrong action type. Now, you need to identify the earliest critical step where you made a mistake, and suggest a correction."
    IN_COMPLETE = "However, your actions did not finish the task, likely more steps are needed. Now, you need to identify the earliest critical step where you made a mistake, and suggest a correction."
    IN_PROGRESS = "However, you took too many steps and yet still did not finish the task. Now, you need to identify the earliest critical step where you made a mistake, and suggest a correction."
    EXCEPTION = "However, your last action is invalid. You should avoid doing that again and try a different action."
    SUCCESS = "Success"
    NONE = "None"

class LiEtAlSession:
    def __init__(self, task_factory: Factory):
        self.task_factory = task_factory
        if not os.path.exists('log/li_et_al_session'):
            os.makedirs('log/li_et_al_session')
        global has_set_up
        if not has_set_up:
            ConsoleLogger.instruction(f'Please set up the browser for the experiment, this includes logging in to necessary sites. Enter any key when you finish.')
            _ = input()
            has_set_up = True

    def train(self, total_iteration = 25, max_steps = 20):
        pass 

    def __is_action_banned(self, action, ban_list):
            r_action = list(filter(lambda x: x.xpath == action.xpath, ban_list))
            r_action = r_action[0] if r_action else None
            if (r_action and isinstance(r_action, InputAction) and isinstance(action, InputAction)) or (isinstance(r_action, ClickAction) and isinstance(action, ClickAction)):
                return True
            return False
                        
    
    def evaluate(self, total_iteration = 25, max_steps = 20, MAX_T = 5):
        ConsoleLogger.info(f'Li et al\'s simulation')

        global simulator
        ConsoleLogger.info(f'Evaluation started for task id {self.task_factory.id}')
        simulator.init(self.task_factory.url)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Iteration', 'Task', 'Chosen Path', 'Status', 'Elapsed Time', 'T'])
        
        for iteration in range(total_iteration):
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} started')
            start_time = time.time()
            simulator.restart(self.task_factory.url)
            time.sleep(2)
            task = self.task_factory.create(simulator)
            ConsoleLogger.info(f'Task: {task}')
            R = [None for _ in range(max_steps)]
            D = [[] for _ in range(max_steps)]
            D_id = [[] for _ in range(max_steps)]
            plan = []
            for T in range(MAX_T):
                short_term_memory = ShortTermMemory(self.task_factory.url, self.task_factory.title, task)
                simulator.restart(self.task_factory.url)
                status = STATUS.NONE
                for step_id in range(max_steps): 
                    if (step_id >= len(plan)):
                        ConsoleLogger.info(f'Generating plan for step {step_id}')
                        if step_id == 0 and T == 0:
                            time.sleep(2)
                        plan += AgentAccess.lis_staged_plan(task, simulator.get_lis_html(), short_term_memory.lis_stringify())                
                    
                    action = plan[step_id]

                    if R[step_id] and not self.__is_action_banned(R[step_id][1], D[step_id]):
                        action = R[step_id][1]
                        plan = plan[0:step_id] + [action] + plan[step_id+1:]
                   
                    prev_dom = simulator.get_lis_html()
                    state = State(simulator.crawl(), None, None, None)
                    disabled_ids = AgentAccess.parse_id_from_lis_actions(D_id[step_id])
                    state.description = simulator.get_lis_html(disabled_ids)
                    ConsoleLogger.info(f'Chosen action: {action.description}')
                    short_term_memory.append(action, state)

                    
                    try:
                        action.perform(simulator)
                    except Exception as e:
                        ConsoleLogger.info(f"Action xpath: {action.xpath}")
                        ConsoleLogger.info(f"Action description: {action.description}")
                        
                        if (isinstance(self.task_factory, MiniwobTaskFatory)):
                            status = STATUS.EXCEPTION
                            break
                        else:
                            ConsoleLogger.instruction(f'Has this xpath correct but browser cause error? (0 = no, 1 = yes): ')
                            is_correct = False if input() == "0" else True
                            if (is_correct):
                                ConsoleLogger.instruction(f'Please perform the action. Enter any key when you finish.')
                                _ = input()
                            if (not is_correct):
                                status = STATUS.EXCEPTION
                                break

                    next_dom = simulator.get_lis_html()
                    if (step_id == max_steps - 1):
                        status = STATUS.IN_PROGRESS
                        break
                    elif (prev_dom == next_dom and not isinstance(action, InputAction)):
                        status = STATUS.NO_CHANGE
                        break
                    
                    if action.description_detail is None:
                        ConsoleLogger.info(f'Action description not found, generating...')
                        action_description = AgentAccess.lis_action_description(prev_dom, action)
                        action.description_detail = action_description
                        short_term_memory.trail[-1]['action'].description_detail = action_description

                    if (isinstance(self.task_factory, MiniwobTaskFatory)):
                        is_done = simulator.has_reach_end()
                        if (is_done):
                            is_successful = simulator.is_successful()
                            if is_successful == None:
                                ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                                is_successful = False if input() == "0" else True
                            if (not is_successful):
                                status = STATUS.FAILED
                            else:
                                status = STATUS.SUCCESS
                            break
                    else:
                        ConsoleLogger.instruction(f'Has this task ended? (0 = no, 1 = yes): ')
                        is_done = False if (input() == "0" or input() == "\n") else True
                        if (is_done):
                            ConsoleLogger.instruction(f'Please help review this task ({task}) (0 = failed, 1 = successful): ')
                            is_successful = False if input() == "0" else True
                            if (not is_successful):
                                status = STATUS.FAILED
                            else:
                                status = STATUS.SUCCESS
                            break
                    
                if status == STATUS.SUCCESS:
                    end_time = time.time()
                    ws.append([iteration, task, short_term_memory.lis_stringify(), 1, end_time - start_time, T+1])
                    break
                elif T == MAX_T - 1:
                    end_time = time.time()
                    ws.append([iteration, task, short_term_memory.lis_stringify(), 0, end_time - start_time, T+1])
                else:
                    critical_step_id, recommended_action = AgentAccess.lis_reflection(self.task_factory.title, short_term_memory.lis_stringify_details(), task, status)
                    critical_step_id = int(critical_step_id)
                    if R[critical_step_id]:
                        D[critical_step_id].append(R[critical_step_id][1])
                    R[critical_step_id] = (plan[critical_step_id], recommended_action)
                    D_id[critical_step_id].append(plan[critical_step_id])
                    for i in range(critical_step_id+1, max_steps):
                        R[i] = None
                        D[i] = []
                        D_id[i] = []
                    plan = plan[0:critical_step_id+1]
            wb.save(f'log/li_et_al_session/{self.task_factory.id}_evaluation.xlsx')
            ConsoleLogger.info(f'Evaluation iteration {iteration + 1} completed')
        ConsoleLogger.info(f'Evaluation ended for url {self.task_factory.url}')
